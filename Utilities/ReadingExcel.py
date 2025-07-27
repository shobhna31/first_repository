import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl  ##module for data driven testing
import XLUtils
###File-->Workbook-->sheet-->rows-->cells(column)

driver=webdriver.Chrome()
driver.maximize_window()
file=r"C:\Users\dell\Desktop\Shobhna-office\SDET\data-driven-testing-sheets\test.xlsx"
# file=os.getcwd() + r"\data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row  ##count number of rows in excel
print("The number of rows are :",rows)
columns=sheet.max_column
print("The number of columns are :",columns)

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end=" ")
    print()


